"""
YouTube Most Subscribed Channels - Data Analysis Project
=========================================================
CSV data ko analyze karke insights aur Excel report generate karta hai.
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import re
import os

# ─────────────────────────────────────────────
# 1. DATA LOADING & CLEANING
# ─────────────────────────────────────────────

def load_and_clean(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip()

    # Subscribers ko numeric banao (footnote references hata ke)
    df["Subscribers (millions)"] = pd.to_numeric(
        df["Subscribers (millions)"], errors="coerce"
    )

    # Language se footnote brackets hataao  [7][8] etc.
    df["Primary language"] = df["Primary language"].apply(
        lambda x: re.sub(r"\[.*?\]", "", str(x)).strip()
    )

    # Rank column add karo
    df.insert(0, "Rank", range(1, len(df) + 1))
    return df


# ─────────────────────────────────────────────
# 2. ANALYSIS FUNCTIONS
# ─────────────────────────────────────────────

def top_n_channels(df, n=10):
    return df.nlargest(n, "Subscribers (millions)")[
        ["Rank", "Name", "Subscribers (millions)", "Category", "Country"]
    ]

def category_stats(df):
    return (
        df.groupby("Category")["Subscribers (millions)"]
        .agg(["count", "sum", "mean", "max"])
        .rename(columns={"count": "Channels", "sum": "Total Subs",
                         "mean": "Avg Subs", "max": "Max Subs"})
        .sort_values("Total Subs", ascending=False)
        .round(2)
    )

def country_stats(df):
    return (
        df.groupby("Country")["Subscribers (millions)"]
        .agg(["count", "sum"])
        .rename(columns={"count": "Channels", "sum": "Total Subs"})
        .sort_values("Total Subs", ascending=False)
        .round(2)
    )

def brand_vs_personal(df):
    return df.groupby("Brand channel")["Subscribers (millions)"].agg(
        ["count", "sum", "mean"]
    ).rename(columns={"count": "Channels", "sum": "Total Subs", "mean": "Avg Subs"}).round(2)

def language_stats(df):
    return (
        df.groupby("Primary language")["Subscribers (millions)"]
        .agg(["count", "sum"])
        .rename(columns={"count": "Channels", "sum": "Total Subs"})
        .sort_values("Channels", ascending=False)
        .round(2)
    )


# ─────────────────────────────────────────────
# 3. EXCEL REPORT GENERATOR
# ─────────────────────────────────────────────

# Style helpers
HDR_FILL   = PatternFill("solid", start_color="1F4E79")   # dark blue
ALT_FILL   = PatternFill("solid", start_color="D6E4F7")   # light blue
GOLD_FILL  = PatternFill("solid", start_color="FFD700")
RED_FILL   = PatternFill("solid", start_color="C00000")
GREEN_FILL = PatternFill("solid", start_color="375623")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=11, name="Arial")
TITLE_FONT = Font(bold=True, color="1F4E79", size=14, name="Arial")
SUB_FONT   = Font(bold=True, color="1F4E79", size=12, name="Arial")
NORMAL     = Font(size=10, name="Arial")

THIN  = Side(style="thin",   color="AAAAAA")
MED   = Side(style="medium", color="1F4E79")
BORDER_THIN  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MED   = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill   = fill
            cell.font   = NORMAL
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")

def write_df_to_sheet(ws, df, start_row=3, title=""):
    # Title
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).font  = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    cols = len(df.columns)
    headers = list(df.columns)

    # Header row
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i).value = h
    style_header_row(ws, start_row, cols)

    # Data rows
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val

    style_data_rows(ws, start_row + 1, start_row + len(df), cols)

    # Auto width
    for i, col in enumerate(headers, 1):
        max_len = max(len(str(col)), df.iloc[:, i-1].astype(str).map(len).max())
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 40)

    return start_row + len(df)


def build_excel_report(df, out_path):
    wb = Workbook()

    # ── Sheet 1: Dashboard Overview ──────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Dashboard"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = "🎬 YouTube Most Subscribed Channels — Analytics Report"
    title_cell.font  = Font(bold=True, color="FFFFFF", size=16, name="Arial")
    title_cell.fill  = HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    kpis = [
        ("Total Channels Tracked", len(df)),
        ("Total Subscribers (M)",  round(df["Subscribers (millions)"].sum(), 1)),
        ("Avg Subs per Channel (M)", round(df["Subscribers (millions)"].mean(), 1)),
        ("Highest Subs (M)",        df["Subscribers (millions)"].max()),
        ("Total Countries",         df["Country"].nunique()),
        ("Total Categories",        df["Category"].nunique()),
    ]
    kpi_colors = ["1F4E79", "375623", "C55A11", "843C0C", "7030A0", "0070C0"]

    ws1.row_dimensions[3].height = 20
    ws1["A3"] = "📌 Key Metrics"
    ws1["A3"].font = SUB_FONT

    for i, ((label, value), color) in enumerate(zip(kpis, kpi_colors), 1):
        col = get_column_letter(i)
        label_cell = ws1[f"{col}4"]
        val_cell   = ws1[f"{col}5"]
        label_cell.value = label
        val_cell.value   = value
        label_cell.fill  = PatternFill("solid", start_color=color)
        val_cell.fill    = PatternFill("solid", start_color=color)
        label_cell.font  = Font(bold=True, color="FFFFFF", size=9, name="Arial")
        val_cell.font    = Font(bold=True, color="FFFFFF", size=14, name="Arial")
        label_cell.alignment = Alignment(horizontal="center", wrap_text=True)
        val_cell.alignment   = Alignment(horizontal="center")
        ws1.column_dimensions[col].width = 22
        ws1.row_dimensions[4].height = 32
        ws1.row_dimensions[5].height = 28

    # Top 10 table in dashboard
    ws1["A7"] = "🏆 Top 10 Channels"
    ws1["A7"].font = SUB_FONT
    ws1.row_dimensions[7].height = 20

    top10 = top_n_channels(df, 10)
    write_df_to_sheet(ws1, top10, start_row=8, title="")

    # Medal colors for top 3
    medal = {9: "FFD700", 10: "C0C0C0", 11: "CD7F32"}
    for r, color in medal.items():
        for c in range(1, len(top10.columns) + 1):
            ws1.cell(row=r, column=c).fill = PatternFill("solid", start_color=color)
            ws1.cell(row=r, column=c).font = Font(bold=True, size=10, name="Arial")

    # ── Sheet 2: All Channels ─────────────────────────────────
    ws2 = wb.create_sheet("📋 All Channels")
    ws2.sheet_view.showGridLines = False
    write_df_to_sheet(ws2, df, start_row=3, title="All Channels — Complete Data")

    # Freeze header
    ws2.freeze_panes = "A4"

    # ── Sheet 3: Category Analysis ────────────────────────────
    ws3 = wb.create_sheet("📂 By Category")
    ws3.sheet_view.showGridLines = False
    cat = category_stats(df).reset_index()
    last_row = write_df_to_sheet(ws3, cat, start_row=3, title="Subscriber Analysis by Category")

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Total Subscribers by Category (M)"
    chart.y_axis.title = "Subscribers (M)"
    chart.x_axis.title = "Category"
    chart.style = 10
    chart.width = 22
    chart.height = 13

    data_ref  = Reference(ws3, min_col=3, max_col=3, min_row=3, max_row=3 + len(cat))
    cats_ref  = Reference(ws3, min_col=1, min_row=4, max_row=3 + len(cat))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws3.add_chart(chart, f"A{last_row + 3}")

    # ── Sheet 4: Country Analysis ─────────────────────────────
    ws4 = wb.create_sheet("🌍 By Country")
    ws4.sheet_view.showGridLines = False
    cntry = country_stats(df).reset_index()
    write_df_to_sheet(ws4, cntry, start_row=3, title="Subscriber Analysis by Country")

    # ── Sheet 5: Brand vs Personal ────────────────────────────
    ws5 = wb.create_sheet("🏷️ Brand vs Personal")
    ws5.sheet_view.showGridLines = False
    bvp = brand_vs_personal(df).reset_index()
    write_df_to_sheet(ws5, bvp, start_row=3, title="Brand Channel vs Personal Channel Comparison")

    # ── Sheet 6: Language Analysis ────────────────────────────
    ws6 = wb.create_sheet("🗣️ By Language")
    ws6.sheet_view.showGridLines = False
    lang = language_stats(df).reset_index()
    write_df_to_sheet(ws6, lang, start_row=3, title="Channels & Subscribers by Language")

    wb.save(out_path)
    print(f"✅ Excel report saved: {out_path}")


# ─────────────────────────────────────────────
# 4. MATPLOTLIB CHARTS (PNG)
# ─────────────────────────────────────────────

def save_charts(df, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    plt.style.use("ggplot")

    # Chart 1: Top 10 bar
    top10 = df.nlargest(10, "Subscribers (millions)")
    fig, ax = plt.subplots(figsize=(12, 6))
    colors = ["#FFD700" if i == 0 else "#C0C0C0" if i == 1 else "#CD7F32"
              if i == 2 else "#1F4E79" for i in range(10)]
    bars = ax.barh(top10["Name"][::-1], top10["Subscribers (millions)"][::-1], color=colors[::-1])
    ax.set_xlabel("Subscribers (Millions)", fontsize=12)
    ax.set_title("Top 10 Most Subscribed YouTube Channels", fontsize=14, fontweight="bold")
    for bar, val in zip(bars, top10["Subscribers (millions)"][::-1]):
        ax.text(bar.get_width() + 1, bar.get_y() + bar.get_height()/2,
                f"{val}M", va="center", fontsize=9)
    plt.tight_layout()
    fig.savefig(f"{out_dir}/top10_channels.png", dpi=150, bbox_inches="tight")
    plt.close()

    # Chart 2: Category pie
    cat = df.groupby("Category")["Subscribers (millions)"].sum().sort_values(ascending=False)
    fig, ax = plt.subplots(figsize=(9, 7))
    wedges, texts, autotexts = ax.pie(
        cat.values, labels=cat.index, autopct="%1.1f%%",
        startangle=140, pctdistance=0.82,
        colors=plt.cm.Set3.colors[:len(cat)]
    )
    ax.set_title("Subscriber Share by Category", fontsize=13, fontweight="bold")
    plt.tight_layout()
    fig.savefig(f"{out_dir}/category_pie.png", dpi=150, bbox_inches="tight")
    plt.close()

    # Chart 3: Country bar (top 8)
    top_countries = df.groupby("Country")["Subscribers (millions)"].sum().nlargest(8)
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(top_countries.index, top_countries.values, color="#1F4E79", edgecolor="white")
    ax.set_ylabel("Total Subscribers (M)")
    ax.set_title("Top Countries by Total Subscribers", fontsize=13, fontweight="bold")
    plt.xticks(rotation=30, ha="right")
    plt.tight_layout()
    fig.savefig(f"{out_dir}/country_bar.png", dpi=150, bbox_inches="tight")
    plt.close()

    print(f"✅ Charts saved in: {out_dir}/")


# ─────────────────────────────────────────────
# 5. MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    CSV_PATH  = "Most_Subscribed_YouTube_Channels_exported.csv"
    XLSX_OUT  = "YouTube_Analysis_Report.xlsx"
    CHART_DIR = "charts"

    print("=" * 55)
    print("  YouTube Channel Data Analysis Project")
    print("=" * 55)

    df = load_and_clean(CSV_PATH)

    # ── Console Summary ──────────────────────────────────────
    print(f"\n📊 Dataset loaded: {len(df)} channels\n")

    print("🏆 Top 5 Channels:")
    print(top_n_channels(df, 5).to_string(index=False))

    print("\n📂 Category Stats:")
    print(category_stats(df).to_string())

    print("\n🌍 Top 5 Countries:")
    print(country_stats(df).head(5).to_string())

    print("\n🏷️  Brand vs Personal:")
    print(brand_vs_personal(df).to_string())

    print("\n🗣️  Language Distribution:")
    print(language_stats(df).to_string())

    # ── Excel Report & Charts ────────────────────────────────
    print("\n⏳ Generating Excel report...")
    build_excel_report(df, XLSX_OUT)

    print("⏳ Generating charts...")
    save_charts(df, CHART_DIR)

    print("\n✅ Project complete!")
    print(f"   📁 Excel  → {XLSX_OUT}")
    print(f"   📁 Charts → {CHART_DIR}/")
